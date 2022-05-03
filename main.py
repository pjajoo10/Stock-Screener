import timeit
from asyncio.windows_events import NULL
from time import sleep
from datetime import date
import pandas as pd
import numpy as np

from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver import FirefoxOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def write_data(data):
    df = pd.DataFrame(data)
    d = date.today().strftime("%d %B, %Y")
    wb = load_workbook(filename="stock_data.xlsx")
    ws = wb.create_sheet(title=str(d))

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    wb.save("stock_data.xlsx")

def screener(indices, latest_data, prev_data):
    s = Service(executable_path="webdrivers\\geckodriver-v0.31.0-win64\\geckodriver.exe")
    o = FirefoxOptions()
    browser = webdriver.Firefox(service=s, options=o)
    #browser.maximize_window()

    data = {
        'SYMBOL': [],
        'NET PROFIT': [],
        'ROCE': [],
        'DEBT/EQUITY': [],
        'DIVIDEND YIELD': [],
        'PRICE-BOOK RATIO': [],
        'PE RATIO': [],
        'VOL CHANGE': []
    }

    start = timeit.default_timer()

    for i, index in enumerate(indices, 1):
        url = "https://www.screener.in/company/"+index+"/consolidated"
        browser.get(url)
        try:
            WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.XPATH, "/html/body/main/div[2]/div[1]/div/h1")))
            data['SYMBOL'].append(index)
            #sleep(.2)
            # net profit for past 3 years
            try:
                net_profit = browser.find_elements(By.XPATH, "//section[@id='profit-loss']/div[@class='responsive-holder fill-card-width']/table/tbody/tr[10]/td")
                if(len(net_profit)>=5):
                    if(int(net_profit[-2].text.replace(',',''))>0 and int(net_profit[-3].text.replace(',',''))>0 and int(net_profit[-4].text.replace(',',''))>0):
                        data['NET PROFIT'].append(True)
                    else:
                        data['NET PROFIT'].append(False)
                else:
                    data['NET PROFIT'].append(NULL)
            except ValueError:
                data['NET PROFIT'].append(NULL)

            # roce > 10%
            try:
                roce = browser.find_element(By.XPATH, "//div[@class='company-ratios']/ul/li[7]/span[2]/span").text
                data['ROCE'].append(float(roce))
            except ValueError:
                data['ROCE'].append(NULL)

            # debt equity < 1.33
            try:
                debt = browser.find_elements(By.XPATH, "//section[@id='balance-sheet']/div[2]/table/tbody/tr[3]/td")
                reserves = browser.find_elements(By.XPATH, "//section[@id='balance-sheet']/div[@class='responsive-holder fill-card-width']/table/tbody/tr[2]/td")
                equity = browser.find_elements(By.XPATH, "//section[@id='balance-sheet']/div[2]/table/tbody/tr[1]/td")
                if(len(debt)>2):
                    de_ratio = int(debt[-2].text.replace(',',''))/(int(reserves[-2].text.replace(',',''))+int(equity[-2].text.replace(',','')))
                    data['DEBT/EQUITY'].append(round(de_ratio, 2))
                else:
                    data['DEBT/EQUITY'].append(NULL)
            except ValueError:
                data['DEBT/EQUITY'].append(NULL)

            # dividend yield > 1%
            try:
                div_yield = browser.find_element(By.XPATH, "//div[@class='company-ratios']/ul/li[6]/span[2]/span").text
                data['DIVIDEND YIELD'].append(float(div_yield))
            except ValueError:
                data['DIVIDEND YIELD'].append(NULL)

            # price to book < 5
            try:
                curr_price = browser.find_element(By.XPATH, "//div[@class='company-ratios']/ul/li[2]/span[2]/span").text.replace(',','')
                book_price = browser.find_element(By.XPATH, "//div[@class='company-ratios']/ul/li[5]/span[2]/span").text.replace(',','')
                if(float(book_price)==0):
                    data['PRICE-BOOK RATIO'].append(NULL)
                else:
                    pb_ratio = float(curr_price)/float(book_price)
                    data['PRICE-BOOK RATIO'].append(round(pb_ratio, 2))
            except ValueError:
                data['PRICE-BOOK RATIO'].append(NULL)

            # price to earnings < 20
            try:
                pe_ratio = browser.find_element(By.XPATH, "//div[@class='company-ratios']/ul/li[4]/span[2]/span").text
                data['PE RATIO'].append(float(pe_ratio))
            except ValueError:
                data['PE RATIO'].append(NULL)

            # volume change from previous day
            curr_vol = latest_data.at[i, 'VOLUME \n(shares)']
            try:
                prev_vol = prev_data[prev_data['SYMBOL \n']==index]['VOLUME \n(shares)'].values[0]
            except KeyError:
                prev_vol = curr_vol
            data['VOL CHANGE'].append(curr_vol/prev_vol)

            sleep(2.8)
        except TimeoutException:
            print("timed out!")
            stop = timeit.default_timer()
            et = stop - start
            print(str(et)+"s")

            browser.quit()
            return(data)

    stop = timeit.default_timer()
    et = stop - start
    print(str(et)+"s")

    browser.quit()
    return(data)


if __name__ == "__main__":
    latest_data = pd.read_csv("market reports\\MW-NIFTY500-MULTICAP-50 25 25-10-Feb-2022.csv")
    prev_data = pd.read_csv("market reports\\MW-NIFTY500-MULTICAP-50 25 25-09-Feb-2022.csv")
    #print(latest_data.head())

    indices = np.array(latest_data['SYMBOL \n'][1:])

    data = screener(indices, latest_data, prev_data)
    #print(data)
    write_data(data)
